import os
import sys
import openpyxl as op

if getattr(sys, 'frozen', False):
    basedir = sys._MEIPASS
else:
    basedir = os.path.dirname(os.path.abspath(__file__))

def excel_append():

    excel_name = "样式文档.xlsx"

    new_excel_name = "新样式文档.xlsx"

    file_path = os.path.join(basedir, excel_name)

    wb = op.load_workbook(file_path)

    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:

        ws = wb[sheet_name]

        insert_data = {"Sheet1":[["林俊杰","男","18"],["邓紫棋","女","20"]],"Sheet2":[["林俊杰","修炼爱情","888W"],["邓紫棋","龙卷风","777W"]]}

        max_row = ws.max_row + 1

        ws.insert_rows(max_row)
        
        for i in range(len(insert_data[sheet_name])):
            for j in range(len(insert_data[sheet_name][i])):
                ws.cell(max_row + i, j + 1, insert_data[sheet_name][i][j])

        wb.save(new_excel_name)

